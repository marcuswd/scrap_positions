import openpyxl
import time
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException

driver = webdriver.Chrome()

sources = []
list_roles = []
positions = []

input_locale = input("Deseja buscar vagas no BR ou Exterior? (br/ext) [Deixe em branco para buscar no Exterior]")

if input_locale == "br":
  sources_path = "br/sources.txt"
  string_search = "%22remoto%22"
else :
  input_locale = "ext"
  sources_path = "ext/sources.txt"
  string_search = "%22remote%22+-%22remote+in+the+US%22"

input_start_date = input("Qual a data de início da busca? (AAAA-MM-DD) [Deixe em branco para buscar vagas dos últimos 7 dias]")

if input_start_date != "":
  start_date = input_start_date
else:
  today = datetime.now().date().strftime('%Y-%m-%d')
  start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')

workbook = openpyxl.Workbook()
sheet = workbook.active

with open(sources_path, 'r', encoding='utf-8') as file:
  for line in file:
    sources.append(line.strip())
    
with open('roles.txt', 'r', encoding='utf-8') as file:
  for line in file:
    list_roles.append(line.strip())
    

for role in list_roles:
  current_tab = workbook.create_sheet(role)
  print(f"Buscando vagas para {role} - Total vacancies = {len(positions)}")
  for source in sources:
    url = f"https://www.google.com/search?q=site:{source}+%22{role}%22+{string_search}+\"+after:{start_date}"
    driver.get(url)
    
    # Wait for the CAPTCHA to be resolved
    while True:
      try:
          # Check for an element that appears only after the CAPTCHA is resolved
          driver.find_element(By.CSS_SELECTOR, '#botstuff')
          break  # CAPTCHA is resolved, exit the loop
      except NoSuchElementException:
          # CAPTCHA is not resolved yet, wait for a while before checking again
          time.sleep(5)
  
    while True:
      print(f"Extracting job listings from {driver.current_url}")
      # Extract job listings from the current page
      result_content = driver.find_elements(By.CSS_SELECTOR, '#search div > div h3')
      if result_content:
        for position in result_content:
          positions.append(position)
          job_title = position.text
          job_url = position.find_element(By.XPATH, '..').get_attribute('href')
          current_tab.append([job_title, job_url])
      
      try:
          # Find and click the "Next" button to go to the next page
          next_button = driver.find_element(By.ID, 'pnnext')
          next_button.click()
          
          # Wait for the next page to load
          time.sleep(2)
          
          # Wait for the CAPTCHA to be resolved again if it appears
          while True:
              try:
                  driver.find_element(By.CSS_SELECTOR, '#botstuff')
                  break
              except NoSuchElementException:
                  time.sleep(5)
      except NoSuchElementException:
          # No more pages left
          break

# Save the workbook
if len(positions) > 0:
  if 'Sheet' in workbook.sheetnames:
    print("Salvando arquivo...")
    workbook.remove(workbook['Sheet'])
    workbook.save(f"vagas_{input_locale}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
else:
  print("Nenhuma vaga encontrada")
  driver.quit()